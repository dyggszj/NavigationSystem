from openpyxl import load_workbook
import numpy as np

def myPrint(name,var):
    try:
        print(name+":"+var)
    except :
        print(name+":"+str(var))
    
def intergrate(arr,padding):
    ret = []
    for x in range(padding,arr.shape[0]-padding):
        temp = []
        temp = np.array(temp)
        for y in range(x-padding,x+padding+1):
            temp = np.append(temp,arr[y])
        ret.append(temp)
    return np.array(ret)
                
def readXlsx(path):
    workbook = load_workbook(path, data_only=True)    #找到需要xlsx文件的位置
    booksheet = workbook.active                 #获取当前活跃的sheet,默认是第一个sheet

    #如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
    # sheets = workbook.get_sheet_names()  # 从名称获取sheet
    # booksheet = workbook.get_sheet_by_name(sheets[0])

    #获取sheet页的行数据
    rows = booksheet.rows
    #获取sheet页的列数据
    # columns = booksheet.columns
    # myPrint("rows",str(rows))
    # myPrint('col',str(columns))
    i = 0
    # 迭代所有的行
    retdata = []
    for row in rows:
        i = i + 1
        line = [col.value for col in row]
        retdata.append(line)
        # cell_data_1 = booksheet.cell(row=i, column=3).value               #获取第i行1 列的数据
        # cell_data_2 = booksheet.cell(row=i, column=4).value               #获取第i行 2 列的数据
        # cell_data_3 = booksheet.cell(row=i, column=8).value                   #获取第i行 3 列的数据
        # cell_data_4 = booksheet.cell(row=i, column=18).value                   #获取第i行 4 列的数据
        # print (cell_data_1, cell_data_2, cell_data_3, cell_data_4)
    return retdata
if __name__ == '__main__':
    arr = [[0,0,0],[1,1,1],[2,2,2],[3,3,3],[4,4,4]]
    arr = np.array(arr)
    x  = intergrate(arr,1)
    print(x)